import openpyxl
from celery import shared_task
from core.models import Customer, Loan
from datetime import datetime


@shared_task
def ingest_customers(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        customer_id, first, last, phone, salary, limit_amt, debt = row

        Customer.objects.update_or_create(
            external_customer_id=customer_id,
            defaults={
                "first_name": first,
                "last_name": last,
                "phone_number": int(phone) if phone else None,
                "age": 30,  # not provided in excel, safe default
                "monthly_salary": salary,
                "approved_limit": limit_amt,
                "current_debt": debt,
            }
        )

@shared_task
def ingest_loans(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        (
            customer_id,
            loan_id,
            amount,
            tenure,
            interest,
            emi,
            emis_paid,
            start_date,
            end_date
        ) = row

        customer = Customer.objects.filter(
            external_customer_id=customer_id
        ).first()

        if not customer:
            continue

        Loan.objects.update_or_create(
            id=loan_id,
            defaults={
                "customer": customer,
                "loan_amount": amount,
                "tenure": tenure,
                "interest_rate": interest,
                "monthly_installment": emi,
                "emis_paid_on_time": emis_paid,
                "start_date": start_date,
                "end_date": end_date,
            }
        )

